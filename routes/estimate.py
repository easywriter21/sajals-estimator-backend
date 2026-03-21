from fastapi import APIRouter
from fastapi.responses import FileResponse
import pandas as pd
from utils.cost_calculator import calculate_cost
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, Reference

router = APIRouter(prefix="/excel")

@router.post("/")
def export_excel(data: dict):

    # 🔹 Step 1: Calculate
    result = calculate_cost(data)
    boq = result["boq"]

    df = pd.DataFrame(boq)

    file_name = "Sajals_Estimator_BOQ.xlsx"

    # 🔹 Step 2: Create Excel
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="BOQ", index=False)

        summary = pd.DataFrame({
            "Parameter": [
                "Total Area (sqft)",
                "Cement (bags)",
                "Steel (kg)",
                "Sand (m3)",
                "Aggregate (m3)",
                "Cost per sqft",
                "Total Cost"
            ],
            "Value": [
                result["total_area_sqft"],
                result["cement_bags"],
                result["steel_kg"],
                result["sand_m3"],
                result["aggregate_m3"],
                result["cost_per_sqft"],
                result["estimated_cost"]
            ]
        })

        summary.to_excel(writer, sheet_name="Summary", index=False)

    # 🔹 Step 3: Styling with openpyxl
    wb = load_workbook(file_name)

    # ---------- BOQ Styling ----------
    ws = wb["BOQ"]

    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Style header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Add borders + center align
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

    # ---------- Add TOTAL Formula ----------
    last_row = ws.max_row + 1
    ws[f"A{last_row}"] = "TOTAL"
    ws[f"E{last_row}"] = f"=SUM(E2:E{last_row-1})"
    ws[f"A{last_row}"].font = Font(bold=True)
    ws[f"E{last_row}"].font = Font(bold=True)

    # ---------- SUMMARY Styling ----------
    ws2 = wb["Summary"]

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in ws2.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # Highlight total cost
    for row in ws2.iter_rows():
        if row[0].value == "Total Cost":
            row[1].font = Font(bold=True, color="FF0000")

    # ---------- PIE CHART ----------
    chart = PieChart()
    chart.title = "Cost Distribution"

    data = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row-1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row-1)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)

    ws.add_chart(chart, "G2")

    wb.save(file_name)

    return FileResponse(
        path=file_name,
        filename=file_name,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
