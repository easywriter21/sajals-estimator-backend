from fastapi import APIRouter
from fastapi.responses import FileResponse
import pandas as pd
from utils.cost_calculator import calculate_cost
from openpyxl import load_workbook
from openpyxl.styles import Font

router = APIRouter(prefix="/excel")

@router.post("/")
def export_excel(data: dict):

    result = calculate_cost(data)
    df = pd.DataFrame(result["boq"])

    file_name = "BOQ.xlsx"

    df.to_excel(file_name, index=False)

    wb = load_workbook(file_name)
    ws = wb.active

    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(file_name)

    return FileResponse(file_name, filename=file_name)
